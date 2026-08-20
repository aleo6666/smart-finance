"""账单批量导入解析器：微信 / 支付宝 / 通用 CSV + Excel(.xlsx)。

对齐旧 Node 后端 server/src/services/import/billParser.js：
- 编码：UTF-8(BOM) → GBK/GB18030 自动探测
- 格式自动识别：微信（微信支付账单+交易时间）、支付宝（支付宝+交易分类）、
  通用（日期列+金额列）
- 统一输出 {type, amount(Decimal), date, category, merchant, description,
  source_type, raw}
- 分类映射表与 Node 完全一致（餐饮美食→餐饮 等）
- Excel 支持合并单元格（openpyxl values_only 天然只取锚点值）、序列号/日期对象转换

重复检测的 calculate_similarity 也移植于此（供 import_batches 服务复用）。
"""

from __future__ import annotations

import csv
import io
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
import re

from app.services.importers.csv_bill import decode_bytes

MONEY = Decimal("0.01")

WECHAT_CATEGORY_MAP = {
    "餐饮美食": "餐饮",
    "食品饮料": "餐饮",
    "交通出行": "交通",
    "交通": "交通",
    "购物消费": "购物",
    "购物": "购物",
    "休闲娱乐": "娱乐",
    "娱乐": "娱乐",
    "居家生活": "居家",
    "生活服务": "居家",
    "家居家装": "居家",
    "医疗健康": "医疗",
    "教育学习": "教育",
    "文化休闲": "教育",
    "运动健身": "运动",
    "旅行交通": "旅行",
    "酒店旅行": "旅行",
    "转账红包": "转账",
    "红包": "转账",
    "转账": "转账",
    "充值缴费": "缴费",
    "数码电器": "数码",
    "美容美发": "美容",
    "其他": "其他",
}

ALIPAY_CATEGORY_MAP = {
    "餐饮美食": "餐饮",
    "交通出行": "交通",
    "服饰装扮": "购物",
    "日用百货": "居家",
    "家居家装": "居家",
    "休闲娱乐": "娱乐",
    "医疗健康": "医疗",
    "教育": "教育",
    "教育培训": "教育",
    "运动户外": "运动",
    "旅行": "旅行",
    "酒店": "旅行",
    "转账红包": "转账",
    "充值缴费": "缴费",
    "手机通讯": "缴费",
    "数码电器": "数码",
    "美容美发": "美容",
    "保险理财": "理财",
    "其他": "其他",
}


class BillParseError(Exception):
    """账单解析失败。"""


# ========= 分类映射 =========


def map_category(category: str, source_type: str) -> str:
    if not category:
        return "其他"
    mapping = (
        WECHAT_CATEGORY_MAP
        if source_type == "wechat"
        else ALIPAY_CATEGORY_MAP
        if source_type == "alipay"
        else {}
    )
    return mapping.get(category) or category or "其他"


# ========= 基础字段解析 =========


def _parse_amount(raw) -> Decimal:
    if raw is None:
        return Decimal("0")
    cleaned = re.sub(r"[^\d.\-]", "", str(raw))
    if not cleaned or cleaned in {".", "-", "-."}:
        return Decimal("0")
    try:
        value = Decimal(cleaned)
    except InvalidOperation:
        return Decimal("0")
    if not value.is_finite():
        return Decimal("0")
    return abs(value).quantize(MONEY, rounding=ROUND_HALF_UP)


def _parse_date(raw) -> str:
    if not raw:
        return ""
    text = str(raw).strip()
    match = re.match(r"^(\d{4})[-/](\d{1,2})[-/](\d{1,2})", text)
    if match:
        year, month, day = match.groups()
        return f"{year}-{month.zfill(2)}-{day.zfill(2)}"
    return text[:10]


def _parse_type(raw_direction, amount: Decimal) -> str:
    direction = str(raw_direction or "").strip()
    if direction in {"收入", "收"}:
        return "income"
    if direction in {"支出", "支"}:
        return "expense"
    # "不计收支" 等中性交易按金额符号推断（金额取绝对值后默认支出）
    return "expense" if amount >= 0 else "income"


def _find_col(row: dict, pattern: str) -> str | None:
    for key in row:
        if re.search(pattern, key, re.IGNORECASE):
            return key
    return None


# ========= 各平台行解析 =========


def _parse_wechat_row(row: dict) -> dict:
    amount = _parse_amount(row.get("金额(元)") or row.get("金额"))
    category = map_category(row.get("交易类型") or row.get("交易分类"), "wechat")
    return {
        "type": _parse_type(row.get("收/支"), amount),
        "amount": amount,
        "date": _parse_date(row.get("交易时间")),
        "category": category,
        "merchant": row.get("交易对方") or "",
        "description": row.get("商品") or row.get("商品说明") or "",
        "source_type": "wechat",
        "raw": dict(row),
    }


def _parse_alipay_row(row: dict) -> dict:
    amount = _parse_amount(row.get("金额") or row.get("金额(元)"))
    category = map_category(row.get("交易分类"), "alipay")
    return {
        "type": _parse_type(row.get("收/支"), amount),
        "amount": amount,
        "date": _parse_date(row.get("交易时间")),
        "category": category,
        "merchant": row.get("交易对方") or "",
        "description": row.get("商品说明") or "",
        "source_type": "alipay",
        "raw": dict(row),
    }


def _parse_generic_row(row: dict) -> dict:
    date_col = _find_col(row, r"日期|date|时间|time")
    amount_col = _find_col(row, r"金额|amount|价格|money")
    category_col = _find_col(row, r"分类|category|类型")
    type_col = _find_col(row, r"收支|类型|type|收|支")
    desc_col = _find_col(row, r"描述|备注|说明|desc|remark|note")
    merchant_col = _find_col(row, r"商家|对方|商户|merchant")

    amount = _parse_amount(row.get(amount_col)) if amount_col else Decimal("0")
    type_ = _parse_type(row.get(type_col), amount) if type_col else "expense"
    return {
        "type": type_,
        "amount": amount,
        "date": _parse_date(row.get(date_col)) if date_col else "",
        "category": map_category(row.get(category_col), "generic") if category_col else "其他",
        "merchant": (row.get(merchant_col) or "") if merchant_col else "",
        "description": (row.get(desc_col) or "") if desc_col else "",
        "source_type": "generic",
        "raw": dict(row),
    }


# ========= 格式识别与 CSV 解析 =========


def _detect_source_type(rows: list[list[str]]) -> str:
    text = "\n".join(" ".join(row) for row in rows[:30])
    if re.search(r"微信支付账单|微信昵称|微信支付", text) and "交易时间" in text:
        return "wechat"
    if "支付宝" in text and "交易时间" in text and "交易分类" in text:
        return "alipay"
    if re.search(r"日期|时间|date|time", text, re.IGNORECASE) and re.search(
        r"金额|amount|价格|money", text, re.IGNORECASE
    ):
        return "generic"
    return "unknown"


def _clean_header(cell: str) -> str:
    return str(cell or "").strip()


def parse_csv_bill_content(content: str, filename: str = "paste.csv") -> dict:
    """解析 CSV 文本为记录列表。返回 {source_type, records, total_count, valid_count}。"""
    reader = csv.reader(io.StringIO(str(content or "")))
    rows = [
        row
        for row in reader
        if any((cell or "").strip() for cell in row)
    ]
    if not rows:
        return {"source_type": "unknown", "records": [], "total_count": 0, "valid_count": 0}

    source_type = _detect_source_type(rows)

    header_idx = -1
    for index, row in enumerate(rows[:50]):
        joined = " ".join(row)
        if re.search(r"交易时间|日期|date|金额|amount", joined, re.IGNORECASE) and len(row) >= 2:
            header_idx = index
            break
    if header_idx < 0:
        header_idx = 0
    headers = [_clean_header(cell) for cell in rows[header_idx]]

    records: list[dict] = []
    for row in rows[header_idx + 1 :]:
        if len(row) < 2:
            continue
        row_dict = {
            headers[index]: row[index] if index < len(row) else ""
            for index in range(len(headers))
        }
        if source_type == "wechat":
            parsed = _parse_wechat_row(row_dict)
        elif source_type == "alipay":
            parsed = _parse_alipay_row(row_dict)
        else:
            parsed = _parse_generic_row(row_dict)
        if parsed["amount"] > 0 and parsed["date"]:
            records.append(parsed)

    return {
        "source_type": source_type,
        "records": records,
        "total_count": len(rows) - header_idx - 1,
        "valid_count": len(records),
    }


# ========= Excel 解析 =========


def _is_metadata_row(cell_texts: list[str]) -> bool:
    joined = " ".join(cell_texts)
    if re.search(
        r"微信支付账单明细|微信昵称|起始时间|终止时间|导出类型|导出时间|共\d+笔记录|收入：|支出：|中性交易：",
        joined,
    ):
        return True
    if re.match(r"^注[：:]", joined):
        return True
    if re.search(r"本明细仅供|本账单中|若交易记录明细", joined):
        return True
    if re.match(r"^[-—]+\s*微信支付", joined):
        return True
    return False


def _is_header_row(cell_texts: list[str]) -> bool:
    non_empty = [text for text in cell_texts if text.strip()]
    if len(non_empty) < 2:
        return False
    if len({text.strip() for text in non_empty}) < 2:
        return False
    joined = " ".join(cell_texts)
    return bool(
        re.search(
            r"日期|时间|date|time|金额|amount|价格|分类|收支|商家|描述|类型|交易",
            joined,
            re.IGNORECASE,
        )
    )


def _parse_excel_date(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        number = float(value)
        if 1 <= number <= 60000:
            try:
                # Excel 序列号：1899-12-30 起的天数（含 1900 假闰年修正）
                return (date(1899, 12, 30) + timedelta(days=int(number))).isoformat()
            except (OverflowError, ValueError):
                return ""
    return _parse_date(value)


def parse_excel_bill(data: bytes, filename: str) -> dict:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover
        raise BillParseError("服务端未安装 Excel 解析依赖（openpyxl）") from exc
    try:
        workbook = load_workbook(io.BytesIO(data), data_only=True)
    except Exception as exc:
        raise BillParseError(f"无法解析 Excel 文件，请确认是 .xlsx 格式") from exc

    sheets = [ws for ws in workbook.worksheets if ws.max_row and ws.max_row >= 2]
    if not sheets:
        return {"source_type": "excel", "records": [], "total_count": 0, "valid_count": 0}
    worksheet = sheets[0]

    all_rows: list[list] = []
    for row in worksheet.iter_rows(values_only=True):
        cells = [cell for cell in row if cell is not None]
        if cells:
            all_rows.append(cells)
    if len(all_rows) < 2:
        return {"source_type": "excel", "records": [], "total_count": 0, "valid_count": 0}

    header_idx = -1
    for index, cells in enumerate(all_rows[:40]):
        texts = [str(cell) for cell in cells]
        if _is_header_row(texts) and not _is_metadata_row(texts):
            header_idx = index
            break
    if header_idx < 0:
        header_idx = 0
    headers = [str(cell).strip() for cell in all_rows[header_idx]]

    data_rows: list[list] = []
    for cells in all_rows[header_idx + 1 :]:
        texts = [str(cell) for cell in cells]
        if all(not text.strip() for text in texts):
            continue
        if _is_metadata_row(texts):
            continue
        data_rows.append(cells)

    header_text = ",".join(headers)
    if re.search(r"收/支", header_text) and "交易对方" in header_text:
        source_type = "wechat"
    elif "交易分类" in header_text:
        source_type = "alipay"
    else:
        source_type = "excel"

    records: list[dict] = []
    for values in data_rows:
        row = {
            headers[index]: values[index] if index < len(values) else ""
            for index in range(len(headers))
        }
        date_col = next(
            (header for header in headers if re.search(r"日期|时间|date|time", header, re.IGNORECASE)),
            None,
        )
        date_value = (
            _parse_excel_date(row.get(date_col))
            if date_col
            else _parse_date(row.get("交易时间") or row.get("date") or "")
        )
        if source_type == "wechat":
            parsed = _parse_wechat_row(row)
            if date_value:
                parsed["date"] = date_value
        elif source_type == "alipay":
            parsed = _parse_alipay_row(row)
            if date_value:
                parsed["date"] = date_value
        else:
            amount = _parse_amount(
                row.get("金额(元)") or row.get("金额") or row.get("amount") or row.get("价格")
            )
            parsed = {
                "type": _parse_type(
                    row.get("收/支") or row.get("收支") or row.get("type") or row.get("类型"),
                    amount,
                ),
                "amount": amount,
                "date": date_value,
                "category": map_category(
                    row.get("交易分类") or row.get("分类") or row.get("category") or row.get("类型"),
                    "generic",
                ),
                "merchant": row.get("交易对方") or row.get("商家") or row.get("对方")
                or row.get("merchant") or "",
                "description": row.get("商品") or row.get("商品说明") or row.get("描述")
                or row.get("备注") or row.get("description") or "",
                "source_type": "excel",
                "raw": dict(row),
            }
        if parsed["amount"] > 0 and parsed["date"]:
            records.append(parsed)

    return {
        "source_type": source_type,
        "records": records,
        "total_count": len(data_rows),
        "valid_count": len(records),
    }


# ========= 统一入口 =========


def parse_bill_file(data: bytes, filename: str) -> dict:
    """根据文件名扩展名 / ZIP 魔数自动选择 Excel 或 CSV 解析器。"""
    lower = (filename or "").lower()
    if lower.endswith((".xlsx", ".xls")) or data[:2] == b"PK":
        return parse_excel_bill(data, filename)
    text = decode_bytes(data)
    return parse_csv_bill_content(text, filename)


# ========= 重复检测相似度 =========


def calculate_similarity(a: dict, b: dict) -> float:
    """计算两条记录的相似度（0-1，1 表示完全相同），对齐 Node 实现。"""
    if not a or not b:
        return 0.0
    # 日期不同直接 0
    if a.get("date") and b.get("date") and a["date"] != b["date"]:
        return 0.0
    # 金额差超过 0.01 直接 0
    if abs(float(a.get("amount") or 0) - float(b.get("amount") or 0)) > 0.01:
        return 0.0

    score = 0.0
    total = 4.0
    if abs(float(a.get("amount") or 0) - float(b.get("amount") or 0)) <= 0.01:
        score += 1.0
    if a.get("category") == b.get("category"):
        score += 1.0

    a_text = f"{a.get('merchant') or ''} {a.get('description') or ''}".lower().strip()
    b_text = f"{b.get('merchant') or ''} {b.get('description') or ''}".lower().strip()
    if a_text and b_text:
        if a_text == b_text:
            score += 2.0
        elif a_text in b_text or b_text in a_text:
            score += 1.5
        else:
            a_set = set(a_text)
            b_set = set(b_text)
            overlap = len(a_set & b_set)
            union = len(a_set | b_set)
            if union > 0:
                score += 2.0 * (overlap / union)
    return min(1.0, score / total)
